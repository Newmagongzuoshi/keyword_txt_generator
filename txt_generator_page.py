import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
from pathlib import Path

from config import LEVEL_ORDER


class TxtGeneratorPage(ttk.Frame):
    def __init__(self, parent, extractor_ref):
        super().__init__(parent)
        self.extractor_ref = extractor_ref
        self.mp4_folder = tk.StringVar()
        self.keyword_file = tk.StringVar()
        self.required_keyword_file = tk.StringVar()
        self.required_keyword_order = tk.StringVar(value="随机")
        self.target_length = tk.StringVar(value="800")
        self.extract_mode = tk.StringVar(value="优先提取")
        self.target_level = tk.StringVar(value="区县级")
        self.smart_mode = tk.BooleanVar(value=True)  # 智能模式默认启用
        self.separator = tk.StringVar(value="，")
        self._building = False
        self._setup_ui()
        # 监听关键词文件路径变化，控制编辑框状态
        self.keyword_file.trace_add("write", self._on_keyword_file_change)
        # 初始化智能模式状态
        self._on_smart_mode_toggle()

    def _setup_ui(self):
        main = ttk.Frame(self, padding=10)
        main.pack(fill=tk.BOTH, expand=True)

        title = ttk.Label(main, text="关键词同名 TXT 生成工具 v1.2.0",
                          font=("Microsoft YaHei", 16, "bold"))
        title.pack(pady=(0, 10))

        row1 = ttk.Frame(main)
        row1.pack(fill=tk.X, pady=3)
        ttk.Label(row1, text="MP4 文件夹：", width=14).pack(side=tk.LEFT)
        ttk.Entry(row1, textvariable=self.mp4_folder).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        ttk.Button(row1, text="选择文件夹", command=self._select_mp4_folder).pack(side=tk.LEFT)

        row2 = ttk.Frame(main)
        row2.pack(fill=tk.X, pady=3)
        ttk.Label(row2, text="关键词 TXT：", width=14).pack(side=tk.LEFT)
        ttk.Entry(row2, textvariable=self.keyword_file).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        ttk.Button(row2, text="选择 TXT", command=self._select_keyword_file).pack(side=tk.LEFT)

        row_req_file = ttk.Frame(main)
        row_req_file.pack(fill=tk.X, pady=3)
        ttk.Label(row_req_file, text="必选关键词 TXT：", width=14).pack(side=tk.LEFT)
        ttk.Entry(row_req_file, textvariable=self.required_keyword_file).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        ttk.Button(row_req_file, text="选择 TXT", command=self._select_required_keyword_file).pack(side=tk.LEFT)

        row_req_order = ttk.Frame(main)
        row_req_order.pack(fill=tk.X, pady=3)
        ttk.Label(row_req_order, text="必选关键词顺序：", width=14).pack(side=tk.LEFT)
        ttk.Combobox(row_req_order, textvariable=self.required_keyword_order,
                     values=["随机", "顺序"], state="readonly", width=12).pack(side=tk.LEFT, padx=5)
        ttk.Label(row_req_order, text="写入必选关键词时的顺序", foreground="gray").pack(side=tk.LEFT)

        row_sep = ttk.Frame(main)
        row_sep.pack(fill=tk.X, pady=3)
        ttk.Label(row_sep, text="分隔符：", width=14).pack(side=tk.LEFT)
        ttk.Combobox(row_sep, textvariable=self.separator,
                     values=["，", "、", ",", "；", "。", "|", " ", "~", "-"],
                     width=6).pack(side=tk.LEFT, padx=5)
        ttk.Label(row_sep, text="生成文案时使用的分隔符", foreground="gray").pack(side=tk.LEFT)

        row3 = ttk.Frame(main)
        row3.pack(fill=tk.X, pady=3)
        ttk.Label(row3, text="目标文本长度：", width=14).pack(side=tk.LEFT)
        ttk.Entry(row3, textvariable=self.target_length, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Label(row3, text="每条文案会在目标长度基础上随机浮动 2%~15%",
                  foreground="gray").pack(side=tk.LEFT)

        row4 = ttk.Frame(main)
        row4.pack(fill=tk.X, pady=3)
        ttk.Label(row4, text="提取模式：", width=14).pack(side=tk.LEFT)
        cb_mode = ttk.Combobox(row4, textvariable=self.extract_mode,
                               values=["优先提取", "强制提取"], state="readonly", width=12)
        cb_mode.pack(side=tk.LEFT, padx=5)
        ttk.Checkbutton(row4, text="智能模式", variable=self.smart_mode,
                        command=self._on_smart_mode_toggle).pack(side=tk.LEFT, padx=(20, 5))
        ttk.Label(row4, text="目标层级：").pack(side=tk.LEFT, padx=(10, 5))
        self.cb_level = ttk.Combobox(row4, textvariable=self.target_level,
                                     values=LEVEL_ORDER, state="readonly", width=12)
        self.cb_level.pack(side=tk.LEFT)

        # 关键词编辑框 - 蓝色边框
        row_keywords = ttk.Frame(main)
        row_keywords.pack(fill=tk.X, pady=3)
        ttk.Label(row_keywords, text="关键词\n（每行一个）：", width=14).pack(side=tk.LEFT, anchor=tk.N)
        self.keywords_text = scrolledtext.ScrolledText(row_keywords, height=4, width=40,
                                                       highlightbackground="#4A90E2", highlightcolor="#4A90E2",
                                                       highlightthickness=2)
        self.keywords_text.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        ttk.Label(row_keywords, text="可不填，可从TXT导入或在此编辑",
                  foreground="gray").pack(side=tk.LEFT)

        # 必选关键词编辑框 - 红色边框
        row5 = ttk.Frame(main)
        row5.pack(fill=tk.X, pady=3)
        ttk.Label(row5, text="必选关键词\n（每行一个）：", width=14).pack(side=tk.LEFT, anchor=tk.N)
        self.required_keywords_text = scrolledtext.ScrolledText(row5, height=4, width=40,
                                                                highlightbackground="#E74C3C", highlightcolor="#E74C3C",
                                                                highlightthickness=2)
        self.required_keywords_text.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        ttk.Label(row5, text="可不填，可从TXT导入或在此编辑",
                  foreground="gray").pack(side=tk.LEFT)

        # 首词编辑框 - 绿色边框
        row6 = ttk.Frame(main)
        row6.pack(fill=tk.X, pady=3)
        ttk.Label(row6, text="首词\n（每行一个）：", width=14).pack(side=tk.LEFT, anchor=tk.N)
        self.first_words_text = scrolledtext.ScrolledText(row6, height=4, width=40,
                                                          highlightbackground="#27AE60", highlightcolor="#27AE60",
                                                          highlightthickness=2)
        self.first_words_text.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        ttk.Label(row6, text="可不填，一行一个首词，\n程序会随机选择一个放在最前面",
                  foreground="gray").pack(side=tk.LEFT)

        btn_row = ttk.Frame(main)
        btn_row.pack(fill=tk.X, pady=10)
        self.start_btn = ttk.Button(btn_row, text="开始生成", command=self._start_generation)
        self.start_btn.pack(side=tk.LEFT, padx=5)
        self.stop_btn = ttk.Button(btn_row, text="停止", command=self._stop_generation, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=5)

        self.progress = ttk.Progressbar(main, mode="determinate")
        self.progress.pack(fill=tk.X, pady=(5, 0))
        self.progress_label = ttk.Label(main, text="")
        self.progress_label.pack()

        self.log_area = scrolledtext.ScrolledText(main, height=14, width=80, state=tk.DISABLED)
        self.log_area.pack(fill=tk.BOTH, expand=True, pady=5)

    def _select_mp4_folder(self):
        folder = filedialog.askdirectory(title="选择 MP4 文件夹")
        if folder:
            self.mp4_folder.set(folder)
            if self.smart_mode.get():
                self._analyze_mp4_files(folder)

    def _on_smart_mode_toggle(self):
        """智能模式切换处理"""
        if self.smart_mode.get():
            # 启用智能模式：禁用手动选择，自动分析
            self.cb_level.config(state="disabled")
            if self.mp4_folder.get():
                self._analyze_mp4_files(self.mp4_folder.get())
        else:
            # 关闭智能模式：启用手动选择
            self.cb_level.config(state="readonly")

    def _analyze_mp4_files(self, folder):
        """智能分析文件夹中的MP4文件，找到每个文件的最低层级地区，设置默认目标层级"""
        import os
        from pathlib import Path
        from collections import Counter

        extractor = self.extractor_ref.get("extractor")
        if not extractor:
            return

        # 扫描MP4文件
        mp4_files = []
        for f in os.listdir(folder):
            if f.lower().endswith('.mp4'):
                mp4_files.append(f)

        if not mp4_files:
            return

        # 统计每个文件的最低层级
        lowest_levels = Counter()
        sample_size = min(50, len(mp4_files))  # 最多分析50个文件作为样本

        for filename in mp4_files[:sample_size]:
            # 获取文件中的所有匹配候选者，找到最低层级
            try:
                candidates = extractor.get_all_candidates(filename)
                if candidates:
                    # 找到所有候选者中的最低层级（数字最大的层级）
                    lowest_level = max(c["level"] for c in candidates)
                    lowest_levels[lowest_level] += 1
            except Exception as e:
                # 如果分析失败，跳过这个文件
                continue

        if not lowest_levels:
            return

        # 找到最常见的最低层级
        most_common_lowest_level = lowest_levels.most_common(1)[0][0]
        level_name_map = {1: "省级", 2: "市级", 3: "区县级", 4: "乡镇街道级", 5: "村社区级"}
        default_level_name = level_name_map.get(most_common_lowest_level, "区县级")

        # 设置默认值
        self.target_level.set(default_level_name)

        # 显示统计信息（如果日志区域已初始化）
        total_analyzed = sum(lowest_levels.values())
        if total_analyzed > 0:
            level_distribution = ", ".join([
                f"{level_name_map[l]}: {count}"
                for l, count in sorted(lowest_levels.items())
            ])
            try:
                self._log(f"智能分析完成：共 {len(mp4_files)} 个MP4文件，"
                         f"分析了 {sample_size} 个样本，"
                         f"各文件最低层级分布：{level_distribution}，"
                         f"智能选择：{default_level_name}")
            except:
                # 如果日志区域还没准备好，静默跳过
                pass


    def _select_keyword_file(self):
        file = filedialog.askopenfilename(
            title="选择关键词 TXT",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")],
        )
        if file:
            self.keyword_file.set(file)

    def _on_keyword_file_change(self, *args):
        path = self.keyword_file.get().strip()
        if path and os.path.isfile(path):
            self.keywords_text.configure(state=tk.DISABLED)
        else:
            self.keywords_text.configure(state=tk.NORMAL)

    def _select_required_keyword_file(self):
        file = filedialog.askopenfilename(
            title="选择必选关键词 TXT",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")],
        )
        if file:
            self.required_keyword_file.set(file)

    def _log(self, msg):
        self.log_area.configure(state=tk.NORMAL)
        self.log_area.insert(tk.END, msg + "\n")
        self.log_area.see(tk.END)
        self.log_area.configure(state=tk.DISABLED)

    def _update_progress(self, current, total):
        self.progress["maximum"] = total
        self.progress["value"] = current
        self.progress_label.configure(text=f"{current}/{total}")

    def _start_generation(self):
        mp4_folder = self.mp4_folder.get().strip()
        keyword_file = self.keyword_file.get().strip()
        keywords_text = self.keywords_text.get("1.0", tk.END)

        if not mp4_folder or not os.path.isdir(mp4_folder):
            messagebox.showwarning("提示", "请先选择有效的 MP4 文件夹")
            return
        has_keyword_file = keyword_file and os.path.isfile(keyword_file)
        has_keywords_text = bool(keywords_text.strip())
        if not has_keyword_file and not has_keywords_text:
            messagebox.showwarning("提示", "请选择关键词 TXT 文件或在关键词编辑框中填写内容")
            return

        try:
            target_length = int(self.target_length.get().strip())
            if target_length < 10:
                messagebox.showwarning("提示", "目标文本长度至少为 10")
                return
        except ValueError:
            messagebox.showwarning("提示", "目标文本长度必须为整数")
            return

        mode = "priority" if self.extract_mode.get() == "优先提取" else "force"
        level_map = {"省级": 1, "市级": 2, "区县级": 3, "乡镇街道级": 4, "村社区级": 5}
        target_level = level_map.get(self.target_level.get(), 3)

        first_words_text = self.first_words_text.get("1.0", tk.END)

        # 检查必选关键词是否超过目标文本长度
        required_keywords_text = self.required_keywords_text.get("1.0", tk.END)
        if required_keywords_text.strip():
            req_lines = [l.strip() for l in required_keywords_text.splitlines() if l.strip()]
            if req_lines:
                total_req_len = sum(len(l) for l in req_lines) + len(req_lines) - 1  # 加上分隔符
                if total_req_len > target_length:
                    messagebox.showwarning(
                        "必选关键词过长",
                        "必选关键词总长度（%d字）超过了目标文本长度（%d字）\n请调整后重试" % (total_req_len, target_length)
                    )
                    self.required_keywords_text.delete("1.0", tk.END)
                    return

        self.start_btn.configure(state=tk.DISABLED)
        self.stop_btn.configure(state=tk.NORMAL)
        self.progress["value"] = 0
        self.progress_label.configure(text="")
        self.log_area.configure(state=tk.NORMAL)
        self.log_area.delete("1.0", tk.END)
        self.log_area.configure(state=tk.DISABLED)

        self._log("正在启动处理...")

        def _log_safe(msg):
            self.after(0, lambda: self._log(msg))

        def _progress_safe(current, total):
            self.after(0, lambda: self._update_progress(current, total))

        self._cancelled = False

        def run():
            try:
                from worker import run_generation
                extractor = self.extractor_ref.get("extractor")
                if extractor is None:
                    _log_safe("错误：地区提取器未初始化")
                    return

                required_keyword_file = self.required_keyword_file.get().strip()
                required_keywords_text = self.required_keywords_text.get("1.0", tk.END)
                required_keyword_order = self.required_keyword_order.get()

                result = run_generation(
                    mp4_folder=mp4_folder,
                    keyword_file=keyword_file,
                    keywords_text=keywords_text,
                    required_keyword_file=required_keyword_file,
                    required_keywords_text=required_keywords_text,
                    required_keyword_order=required_keyword_order,
                    first_words_text=first_words_text,
                    separator=self.separator.get(),
                    target_length=target_length,
                    mode=mode,
                    target_level=target_level,
                    extractor=extractor,
                    log_callback=_log_safe,
                    progress_callback=_progress_safe,
                    smart_mode=self.smart_mode.get(),
                )
                _log_safe("")
                if result.get("success"):
                    _log_safe(result["summary"])
                    # 弹出地址映射对话框
                    region_map = result.get("region_origin_map", {})
                    if region_map:
                        self.after(0, lambda: self._show_region_mapping(region_map))
                else:
                    _log_safe(f"错误：{result.get('error', '未知错误')}")
            except Exception as e:
                _log_safe(f"发生异常：{e}")
            finally:
                self.after(0, self._on_done)

        threading.Thread(target=run, daemon=True).start()

    def _show_region_mapping(self, region_origin_map):
        """弹出地址映射对话框，默认全屏，表格展示，标注降级和可信度"""
        if not region_origin_map:
            return
        dialog = tk.Toplevel(self)
        dialog.title("地址提取映射")
        dialog.configure(bg="#f5f5f5")
        dialog.geometry("960x640")
        dialog.minsize(600, 400)
        dialog.resizable(True, True)

        # 构建行数据：每个地区一行，原地址横向展开为多列
        row_data = []
        max_cols = 0
        for region_name, entries in region_origin_map.items():
            originals = [e["original"] for e in entries]
            fallback_count = sum(1 for e in entries if e["is_fallback"])
            total = len(entries)
            is_degraded = fallback_count > 0
            if fallback_count == 0:
                credibility = "可信"
            elif fallback_count < total:
                credibility = "部分可信"
            else:
                credibility = "低可信"
            row_data.append({
                "region": region_name, "originals": originals,
                "total": total, "credibility": credibility,
                "is_degraded": is_degraded,
            })
            max_cols = max(max_cols, total)

        # 降级优先，文件数从少到多
        row_data.sort(key=lambda r: (not r["is_degraded"], r["total"], r["region"]))

        # 表格容器
        table_frame = tk.Frame(dialog, bg="white", highlightbackground="#ddd", highlightthickness=1)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # 动态列名
        columns = ("序号", "可信度", "提取地址", "文件数") + tuple("原地址%d" % (i+1) for i in range(max_cols))
        tree = ttk.Treeview(table_frame, columns=columns, show="headings", selectmode="extended")
        tree.heading("序号", text="#", anchor=tk.CENTER)
        tree.heading("可信度", text="可信度", anchor=tk.CENTER)
        tree.heading("提取地址", text="提取地址", anchor=tk.W)
        tree.heading("文件数", text="文件数", anchor=tk.CENTER)
        for i in range(max_cols):
            tree.heading("原地址%d" % (i+1), text="原地址%d" % (i+1), anchor=tk.W)

        tree.column("序号", width=40, anchor=tk.CENTER, stretch=False)
        tree.column("可信度", width=80, anchor=tk.CENTER, stretch=False)
        tree.column("提取地址", width=130, anchor=tk.W)
        tree.column("文件数", width=55, anchor=tk.CENTER, stretch=False)
        for i in range(max_cols):
            tree.column("原地址%d" % (i+1), width=260, anchor=tk.W)

        # 行颜色
        tree.tag_configure("degraded", background="#fff3cd")
        tree.tag_configure("normal_even", background="#f8f9fa")
        tree.tag_configure("normal_odd", background="white")
        tree.tag_configure("trusted", foreground="#27ae60")
        tree.tag_configure("partial", foreground="#e67e22")
        tree.tag_configure("low", foreground="#e74c3c")

        cred_color_map = {"可信": "trusted", "部分可信": "partial", "低可信": "low"}

        for i, rd in enumerate(row_data):
            vals = [i + 1, rd["credibility"], rd["region"], rd["total"]]
            vals += rd["originals"] + [""] * (max_cols - len(rd["originals"]))
            tag = "degraded" if rd["is_degraded"] else ("normal_even" if i % 2 == 0 else "normal_odd")
            cred_tag = cred_color_map.get(rd["credibility"], "trusted")
            tree.insert("", tk.END, values=vals, tags=(tag, cred_tag))

        # 统计
        total_regions = len(row_data)
        total_files = sum(r["total"] for r in row_data)
        degraded_regions = sum(1 for r in row_data if r["is_degraded"])

        # 标题栏
        header = tk.Frame(dialog, bg="#2c3e50", height=44)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        tk.Label(header, text="  地区提取映射结果", font=("Microsoft YaHei", 13, "bold"),
                 bg="#2c3e50", fg="white").pack(side=tk.LEFT, pady=10)

        # 统计 + 按钮
        toolbar = tk.Frame(dialog, bg="#ecf0f1")
        toolbar.pack(fill=tk.X, padx=10, pady=(8, 0))
        stats_text = "共 %d 个地区，%d 个文件" % (total_regions, total_files)
        if degraded_regions > 0:
            stats_text += "    |    %d 个地区涉及降级" % degraded_regions
        tk.Label(toolbar, text=stats_text,
                 font=("Microsoft YaHei", 10, "bold"), bg="#ecf0f1", fg="#2c3e50").pack(side=tk.LEFT, pady=6)

        # 滚动条
        vsb = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 图例
        legend = tk.Frame(dialog, bg="#f5f5f5")
        legend.pack(fill=tk.X, padx=15, pady=(2, 0))
        tk.Label(legend, text="可信", fg="#27ae60", font=("Microsoft YaHei", 9), bg="#f5f5f5").pack(side=tk.LEFT, padx=4)
        tk.Label(legend, text="部分可信", fg="#e67e22", font=("Microsoft YaHei", 9), bg="#f5f5f5").pack(side=tk.LEFT, padx=4)
        tk.Label(legend, text="低可信", fg="#e74c3c", font=("Microsoft YaHei", 9), bg="#f5f5f5").pack(side=tk.LEFT, padx=4)
        tk.Label(legend, text="   黄色底色 = 涉及降级", bg="#fff3cd", fg="#856404",
                 font=("Microsoft YaHei", 9)).pack(side=tk.LEFT, padx=10)

        # 底部
        bottom = tk.Frame(dialog, bg="#f5f5f5")
        bottom.pack(fill=tk.X, padx=10, pady=(10, 15))
        ttk.Button(bottom, text="关闭", command=dialog.destroy).pack(side=tk.RIGHT, padx=6, ipadx=8, ipady=4)
        save_btn = tk.Button(bottom, text="  另存为 Excel  ", font=("Microsoft YaHei", 11, "bold"),
                             bg="#27ae60", fg="white", relief=tk.FLAT, cursor="hand2",
                             command=lambda: self._save_mapping(tree, max_cols),
                             activebackground="#219a52", activeforeground="white",
                             padx=16, pady=6)
        save_btn.pack(side=tk.RIGHT, padx=6)

        # 默认最大化打开
        dialog.after(100, lambda: dialog.state("zoomed"))

    def _save_mapping(self, tree, max_cols):
        """保存映射表格为 Excel 文件，默认文件名用视频文件夹名"""
        folder_name = Path(self.mp4_folder.get()).name if self.mp4_folder.get() else "地址映射"
        default_name = "%s-地址映射" % folder_name
        file = filedialog.asksaveasfilename(
            title="保存地址映射",
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
        )
        if not file:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "地址映射"

            # 表头
            headers = ["序号", "可信度", "提取地址", "文件数"] + ["原地址%d" % (i+1) for i in range(max_cols)]
            header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            header_font = Font(name="Microsoft YaHei", size=10, bold=True, color="ffffff")
            thin_border = Border(
                left=Side(style="thin", color="cccccc"),
                right=Side(style="thin", color="cccccc"),
                top=Side(style="thin", color="cccccc"),
                bottom=Side(style="thin", color="cccccc"),
            )

            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            # 数据行
            degrade_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
            even_fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
            cred_fonts = {
                "可信": Font(name="Microsoft YaHei", size=10, color="27ae60"),
                "部分可信": Font(name="Microsoft YaHei", size=10, color="e67e22"),
                "低可信": Font(name="Microsoft YaHei", size=10, color="e74c3c"),
            }
            normal_font = Font(name="Microsoft YaHei", size=10)

            for ri, item in enumerate(tree.get_children()):
                vals = tree.item(item, "values")
                tags = tree.item(item, "tags")
                degraded = "degraded" in tags
                row_fill = degrade_fill if degraded else (even_fill if ri % 2 == 0 else None)
                cred_val = vals[1] if len(vals) > 1 else ""

                for ci, val in enumerate(vals):
                    cell = ws.cell(row=ri + 2, column=ci + 1, value=val)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center" if ci in (0, 1, 3) else "left", vertical="center")
                    if row_fill:
                        cell.fill = row_fill
                    if ci == 1 and cred_val in cred_fonts:
                        cell.font = cred_fonts[cred_val]
                    else:
                        cell.font = normal_font

            # 列宽
            ws.column_dimensions["A"].width = 6
            ws.column_dimensions["B"].width = 10
            ws.column_dimensions["C"].width = 16
            ws.column_dimensions["D"].width = 8
            for i in range(max_cols):
                col_letter = openpyxl.utils.get_column_letter(5 + i)
                ws.column_dimensions[col_letter].width = 32

            # 冻结表头
            ws.freeze_panes = "A2"

            wb.save(file)
            messagebox.showinfo("提示", "已保存到：%s" % file)
        except ImportError:
            messagebox.showwarning("提示", "需要 openpyxl 库，请执行 pip install openpyxl")

    def _stop_generation(self):
        self._cancelled = True
        self._log("用户取消了操作（正在处理的任务会继续完成）")

    def _on_done(self):
        self.start_btn.configure(state=tk.NORMAL)
        self.stop_btn.configure(state=tk.DISABLED)
