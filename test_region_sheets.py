"""从行政区划Excel读取各省数据，生成地区提取报告"""
import os
import sys
from pathlib import Path
from collections import OrderedDict

import openpyxl

from region_db import RegionDatabase
from region_extractor import RegionExtractor

EXCEL_PATH = r"C:\Users\Administrator\Desktop\2023-03-17最新行政区划（呕心沥血版）.xlsx"
REPORT_PATH = Path(__file__).parent / "region_test_report.txt"

# 模拟关键词（用于生成报告中的示例文本）
DEMO_KEYWORDS = ["厂家直销", "工厂批发", "源头厂家", "一件代发", "加工定制", "按需定制"]


def build_full_path(result):
    """从提取结果构建完整行政路径，如 遵义市正安县"""
    region = result.get("region")
    if not region:
        return result["name"]

    chain = region.get("parent_chain", [])
    parts = []
    for p in chain:
        pname = p.get("name", p.get("short_name", ""))
        if pname and pname not in parts:
            parts.append(pname)

    # 最末级用匹配到的名称
    cur_name = region.get("name", result["name"])
    if cur_name and cur_name not in parts:
        parts.append(cur_name)

    return "".join(parts)


def generate_demo_text(area_name, full_path):
    """生成报告中的示例文本"""
    phrases = []
    # 用完整路径+关键词
    phrases.append(full_path + DEMO_KEYWORDS[0])
    phrases.append(DEMO_KEYWORDS[1])
    # 用简称+关键词
    phrases.append(area_name + DEMO_KEYWORDS[1])
    return "，".join(phrases)


def load_extractor():
    db_path = Path(__file__).parent / "region_db.json"
    db = RegionDatabase.load(db_path)
    return RegionExtractor.load(db)


def main():
    print("读取Excel和地区数据库...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    extractor = load_extractor()

    skip_sheets = {"2023-03-17行政区划"}

    # 收集所有sheet的数据
    sheet_data = OrderedDict()
    total_entries = 0
    total_matched = 0
    total_fallback = 0

    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue

        ws = wb[sheet_name]
        entries = []
        for row in ws.iter_rows(values_only=True):
            if row and row[0]:
                val = str(row[0]).strip()
                if val:
                    entries.append(val)

        if not entries:
            continue

        results = []
        matched = 0
        fallback = 0

        for entry in entries:
            filename = f"{entry}.mp4"
            result = extractor.extract(filename, "lowest", 3)
            full_path = build_full_path(result)
            is_fb = result.get("is_fallback", False)

            if is_fb:
                fallback += 1
            else:
                matched += 1

            results.append({
                "entry": entry,
                "extracted": result["name"],
                "full_path": full_path,
                "level": result["level_name"],
                "is_fallback": is_fb,
                "demo_text": generate_demo_text(result["name"], full_path),
            })

        sheet_data[sheet_name] = {
            "count": len(entries),
            "matched": matched,
            "fallback": fallback,
            "results": results,
        }
        total_entries += len(entries)
        total_matched += matched
        total_fallback += fallback

    wb.close()

    # 按条目数从少到多排序
    sorted_sheets = sorted(sheet_data.items(), key=lambda x: x[1]["count"])

    # 生成报告
    lines = []
    lines.append("=" * 70)
    lines.append("地区提取单元测试报告")
    lines.append(f"数据源: 2023-03-17最新行政区划（呕心沥血版）.xlsx")
    lines.append(f"总Sheet数: {len(sorted_sheets)}")
    lines.append(f"总条目数: {total_entries}")
    lines.append(f"匹配成功: {total_matched}")
    lines.append(f"兜底: {total_fallback}")
    lines.append(f"匹配率: {total_matched / total_entries * 100:.1f}%")
    lines.append("=" * 70)
    lines.append("")

    # 汇总表
    lines.append(f"{'Sheet':<16} {'条目':>5} {'匹配':>5} {'兜底':>5} {'匹配率':>7}")
    lines.append("-" * 44)
    for name, data in sorted_sheets:
        rate = data["matched"] / data["count"] * 100 if data["count"] > 0 else 0
        lines.append(f"{name:<16} {data['count']:>5} {data['matched']:>5} "
                     f"{data['fallback']:>5} {rate:>6.1f}%")
    lines.append("-" * 44)
    total_rate = total_matched / total_entries * 100 if total_entries > 0 else 0
    lines.append(f"{'合计':<16} {total_entries:>5} {total_matched:>5} "
                 f"{total_fallback:>5} {total_rate:>6.1f}%")
    lines.append("")

    # 每个Sheet的详细结果
    for sheet_name, data in sorted_sheets:
        lines.append("=" * 70)
        lines.append(f"【{sheet_name}】 共 {data['count']} 条")
        lines.append("=" * 70)

        for i, r in enumerate(data["results"], 1):
            flag = "" if not r["is_fallback"] else " [兜底]"
            lines.append(f"  {i}. {r['extracted']}：{r['demo_text']}{flag}")

        lines.append("")

    # 写入文件
    report_text = "\n".join(lines)
    REPORT_PATH.write_text(report_text, encoding="utf-8")
    print(f"报告已生成: {REPORT_PATH}")
    print(f"总条目: {total_entries}, 匹配: {total_matched}, 兜底: {total_fallback}")
    print(f"匹配率: {total_rate:.1f}%")

    # 同时输出到控制台（前几行）
    print("\n--- 报告摘要（前40行）---")
    for line in lines[:40]:
        print(line)


if __name__ == "__main__":
    main()
